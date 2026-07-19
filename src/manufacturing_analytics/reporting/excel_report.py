"""
excel_report.py
================
Builds a polished, multi-sheet Excel KPI report using openpyxl:
    * Executive Summary  — headline KPIs with conditional (RAG) formatting
    * Trend data          — monthly KPI table backing the summary
    * Factory / Line / Machine breakdowns
    * Quality & Pareto     — defect breakdown table
    * Raw data appendix (optional)

Openpyxl is used directly (rather than ``DataFrame.to_excel`` alone) so we
can apply professional formatting: bold headers, frozen panes, autosized
columns, number formats, and RAG conditional formatting on KPI cells.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from manufacturing_analytics.utils.logger import get_logger

logger = get_logger(__name__)

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
TITLE_FONT = Font(bold=True, size=16, name="Calibri", color="1F4E78")
SUBTITLE_FONT = Font(italic=True, size=10, name="Calibri", color="7F7F7F")
THIN_BORDER = Border(*(Side(style="thin", color="D9D9D9"),) * 4)

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
AMBER_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


class ExcelReportBuilder:
    """Builds a formatted multi-sheet Excel workbook summarizing KPIs."""

    def __init__(self, kpi_targets: dict[str, Any] | None = None):
        self.targets = kpi_targets or {}
        self.wb = Workbook()
        # Remove the default sheet; we add named sheets explicitly.
        self.wb.remove(self.wb.active)

    # ------------------------------------------------------------------ #
    # Public building blocks
    # ------------------------------------------------------------------ #
    def add_title_sheet(self, title: str, subtitle: str, period_label: str) -> None:
        """Add a cover sheet with report title and period."""
        ws = self.wb.create_sheet("Cover", 0)
        ws.sheet_view.showGridLines = False
        ws["B2"] = title
        ws["B2"].font = TITLE_FONT
        ws["B3"] = subtitle
        ws["B3"].font = SUBTITLE_FONT
        ws["B5"] = "Reporting Period:"
        ws["B5"].font = Font(bold=True)
        ws["C5"] = period_label
        ws["B6"] = "Generated:"
        ws["B6"].font = Font(bold=True)
        ws["C6"] = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M UTC")
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 40

    def add_kpi_summary_sheet(self, summary: dict[str, float]) -> None:
        """Add an Executive Summary sheet with headline KPI cards + RAG formatting."""
        ws = self.wb.create_sheet("Executive Summary")
        ws.sheet_view.showGridLines = False
        ws["B2"] = "Executive KPI Summary"
        ws["B2"].font = TITLE_FONT

        headers = ["KPI", "Value", "Target", "Status"]
        start_row = 4
        for c, h in enumerate(headers, start=2):
            cell = ws.cell(row=start_row, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        target_lookup = {
            "OEE": self.targets.get("oee_target", 0.75),
            "Availability": self.targets.get("availability_target", 0.90),
            "Performance": self.targets.get("performance_target", 0.90),
            "Quality": self.targets.get("quality_target", 0.98),
            "Scrap Rate": self.targets.get("scrap_rate_max", 0.03),
            "Defect Rate": self.targets.get("defect_rate_max", 0.03),
        }
        lower_is_better = {"Scrap Rate", "Defect Rate"}

        row = start_row + 1
        for kpi_name, value in summary.items():
            target = target_lookup.get(kpi_name)
            ws.cell(row=row, column=2, value=kpi_name).border = THIN_BORDER
            val_cell = ws.cell(row=row, column=3, value=round(value, 4))
            val_cell.number_format = "0.0%"
            val_cell.border = THIN_BORDER
            if target is not None:
                tgt_cell = ws.cell(row=row, column=4, value=target)
                tgt_cell.number_format = "0.0%"
                tgt_cell.border = THIN_BORDER

                if kpi_name in lower_is_better:
                    status = "Green" if value <= target else ("Amber" if value <= target * 1.2 else "Red")
                else:
                    status = "Green" if value >= target else ("Amber" if value >= target * 0.9 else "Red")
                status_cell = ws.cell(row=row, column=5, value=status)
                status_cell.fill = {"Green": GREEN_FILL, "Amber": AMBER_FILL, "Red": RED_FILL}[status]
                status_cell.alignment = Alignment(horizontal="center")
                status_cell.border = THIN_BORDER
            row += 1

        for col, width in zip("BCDE", [22, 12, 12, 12]):
            ws.column_dimensions[col].width = width

    def add_dataframe_sheet(
        self,
        df: pd.DataFrame,
        sheet_name: str,
        percent_cols: list[str] | None = None,
        freeze_header: bool = True,
    ) -> None:
        """Add a DataFrame as a clean, formatted worksheet (auto-fit columns, formatted header)."""
        ws = self.wb.create_sheet(sheet_name[:31])
        percent_cols = percent_cols or []

        # Header row
        for c, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=c, value=str(col_name))
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for r, row_data in enumerate(df.itertuples(index=False), start=2):
            for c, value in enumerate(row_data, start=1):
                cell = ws.cell(row=r, column=c, value=value)
                cell.border = THIN_BORDER
                col_name = df.columns[c - 1]
                if col_name in percent_cols:
                    cell.number_format = "0.0%"

        self._autofit_columns(ws, df)
        if freeze_header:
            ws.freeze_panes = "A2"
        logger.info("Added Excel sheet '{}' with {} rows.", sheet_name, len(df))

    # ------------------------------------------------------------------ #
    # Save
    # ------------------------------------------------------------------ #
    def save(self, out_path: str | Path) -> Path:
        out_path = Path(out_path)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(out_path)
        logger.success("Excel report saved -> {}", out_path)
        return out_path

    # ------------------------------------------------------------------ #
    # Internal helpers
    # ------------------------------------------------------------------ #
    @staticmethod
    def _autofit_columns(ws: Worksheet, df: pd.DataFrame, max_width: int = 40) -> None:
        for i, col in enumerate(df.columns, start=1):
            max_len = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str).head(200)])
            ws.column_dimensions[get_column_letter(i)].width = min(max_len + 3, max_width)
